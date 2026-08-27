from __future__ import annotations

import csv
import asyncio
import io
import json
import logging
import re
from contextlib import asynccontextmanager
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import BackgroundTasks, FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from .config import get_settings
from .documents import DOCUMENT_ROLES, MAX_STANDARD_FILES, parse_standard_document
from .domain import ModelSettings
from .miner import SpecGapMiner
from .provider import MissingApiKeyError, ProviderError, QianfanProvider
from .service import LabelSpecService
from .store import MAX_RUN_CONCURRENCY, MAX_TRACE_REPLICAS, DatasetDeleteError, StandardDeleteError, Store
from .validator import validate_standard
from .yaml_io import standard_to_yaml_files
from .taxonomy import parse_compiled_standard

app_settings = get_settings()
logging.basicConfig(level=app_settings.labelspec_log_level)
store = Store(app_settings.database_path)
provider = QianfanProvider(app_settings)
service = LabelSpecService(store, provider)
miner = SpecGapMiner(provider, store)


@asynccontextmanager
async def lifespan(_: FastAPI):
    store.initialize()
    yield


app = FastAPI(
    title="LabelSpec API",
    version="0.2.0",
    description="Compile business standards into executable labeling rules.",
    lifespan=lifespan,
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=app_settings.cors_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class CompileRequest(BaseModel):
    name: str = Field(min_length=1, max_length=100)
    source_markdown: str = Field(min_length=20)


class RunRequest(BaseModel):
    dataset_id: str
    standard_id: str
    concurrency: int = Field(default=1, ge=1, le=MAX_RUN_CONCURRENCY)
    trace_replicas: int = Field(default=3, ge=1, le=MAX_TRACE_REPLICAS)


class ReviewRequest(BaseModel):
    human_label: str = Field(min_length=1)
    note: str = ""
    reason_code: str = "MANUAL_REVIEW"
    rule_feedback: List[Dict[str, Any]] = Field(default_factory=list)


class RuleRevisionRequest(BaseModel):
    rule_id: str
    new_rule: Dict[str, Any]
    reason: str = Field(min_length=2)
    related_case_ids: List[str] = Field(default_factory=list)
    suggestion_id: Optional[str] = None


class RulePatchRequest(BaseModel):
    standard_id: str
    payload: Dict[str, Any]
    related_feedback_ids: List[str] = Field(default_factory=list)
    source_run_id: Optional[str] = None


class RulePatchEditRequest(BaseModel):
    payload: Dict[str, Any]


class ManualVersionRequest(BaseModel):
    compiled: Dict[str, Any]
    reason: str = Field(min_length=1, max_length=500)
    resolve_conflicts: bool = False


class ImpactRunRequest(BaseModel):
    source_run_id: str
    target_standard_id: str
    rule_id: str
    labels: List[str] = Field(default_factory=list)
    concurrency: Optional[int] = Field(default=None, ge=1, le=MAX_RUN_CONCURRENCY)
    trace_replicas: Optional[int] = Field(default=None, ge=1, le=MAX_TRACE_REPLICAS)


def _handle_error(exc: Exception) -> HTTPException:
    if isinstance(exc, MissingApiKeyError):
        return HTTPException(status_code=412, detail=str(exc))
    if isinstance(exc, ProviderError):
        return HTTPException(status_code=502, detail=str(exc))
    if isinstance(exc, StandardDeleteError):
        return HTTPException(status_code=409, detail=str(exc))
    if isinstance(exc, DatasetDeleteError):
        return HTTPException(status_code=409, detail=str(exc))
    if isinstance(exc, KeyError):
        return HTTPException(status_code=404, detail=str(exc).strip("'"))
    if isinstance(exc, ValueError):
        return HTTPException(status_code=422, detail=str(exc))
    return HTTPException(status_code=500, detail=str(exc))


def _conflict_source_excerpts(conflict: Dict[str, Any]) -> List[Dict[str, str]]:
    """Return only the matching rule blocks, never the complete source document."""
    names = [part.rsplit("/", 1)[-1].strip() for part in conflict.get("entity_key", "").split(" | ")]
    excerpts: List[Dict[str, str]] = []
    seen = set()
    for ref in conflict.get("source_refs", []):
        document_id = ref.get("document_id")
        if not document_id or document_id in seen:
            continue
        seen.add(document_id)
        try:
            source = store.get_source_document(document_id)
        except KeyError:
            continue
        text = source.get("extracted_text", "")
        lines = text.splitlines()
        starts = [index for index, line in enumerate(lines) if re.match(r"^#{3,6}\s+", line)]
        starts.append(len(lines))
        for start, end in zip(starts, starts[1:]):
            block = "\n".join(lines[start:end]).strip()
            if not block:
                continue
            matched = sum(1 for name in names if name and name in block)
            if matched >= min(2, len(names)):
                excerpts.append({
                    "filename": source["filename"],
                    "locator": ref.get("locator", ""),
                    "excerpt": block[:6000],
                })
    return excerpts


@app.get("/api/health")
def health() -> Dict[str, Any]:
    return {
        "status": "ok",
        "version": "0.2.0",
        "provider": "qianfan",
        "api_key_configured": provider.configured,
        "database": str(app_settings.database_path),
    }


@app.get("/api/settings")
def get_model_settings() -> Dict[str, Any]:
    return {"models": store.get_settings().model_dump(), "api_key_configured": provider.configured}


@app.put("/api/settings")
def put_model_settings(payload: ModelSettings) -> Dict[str, Any]:
    return {"models": store.save_settings(payload).model_dump(), "api_key_configured": provider.configured}


@app.get("/api/models")
async def models() -> Dict[str, Any]:
    try:
        return {"data": await provider.list_models()}
    except Exception as exc:
        raise _handle_error(exc) from exc


@app.get("/api/demo")
def demo_content() -> Dict[str, str]:
    demo_dir = Path(__file__).parent / "demo"
    template = Path(__file__).parent / "templates" / "standard-template.txt"
    return {
        "standard_markdown": (demo_dir / "standard.md").read_text(encoding="utf-8"),
        "dataset_csv": (demo_dir / "data.csv").read_text(encoding="utf-8"),
        "standard_template": template.read_text(encoding="utf-8"),
    }


@app.get("/api/standards/template")
def standard_template() -> Response:
    template = Path(__file__).parent / "templates" / "standard-template.txt"
    return Response(content=template.read_text(encoding="utf-8"), media_type="text/plain; charset=utf-8", headers={"Content-Disposition": 'attachment; filename="standard-template.txt"'})


@app.post("/api/demo/dataset")
def create_demo_dataset() -> Dict[str, Any]:
    path = Path(__file__).parent / "demo" / "data.csv"
    with path.open(encoding="utf-8-sig", newline="") as stream:
        items = list(csv.DictReader(stream))
    return store.create_dataset(store.next_dataset_name("金融与汽车演示数据"), path.name, items)


@app.post("/api/standards/compile")
async def compile_endpoint(payload: CompileRequest) -> Dict[str, Any]:
    try:
        return await service.compile(payload.name, payload.source_markdown)
    except Exception as exc:
        raise _handle_error(exc) from exc


@app.post("/api/standards/compile-files")
async def compile_files_endpoint(
    name: str = Form(..., min_length=1, max_length=100),
    files: List[UploadFile] = File(...),
    base_standard_id: Optional[str] = Form(None),
    roles: List[str] = Form(default=[]),
) -> Dict[str, Any]:
    try:
        if not files or len(files) > MAX_STANDARD_FILES:
            raise ValueError(f"每次需要上传 1 到 {MAX_STANDARD_FILES} 份标准文档")
        if roles and len(roles) != len(files):
            raise ValueError("文档角色数量必须与上传文件数量一致")
        if any(role not in DOCUMENT_ROLES for role in roles):
            raise ValueError("文档角色只能是 auto、definition、boundary 或 priority")
        documents = []
        for index, upload in enumerate(files):
            documents.append(
                parse_standard_document(
                    upload.filename or "standard.txt",
                    upload.content_type or "application/octet-stream",
                    await upload.read(),
                    roles[index] if roles else "auto",
                )
            )
        return await service.compile_documents(name, documents, base_standard_id)
    except Exception as exc:
        raise _handle_error(exc) from exc


@app.get("/api/standards")
def standards() -> List[Dict[str, Any]]:
    return store.list_standards()


@app.get("/api/standards/{standard_id}")
def standard_detail(standard_id: str) -> Dict[str, Any]:
    try:
        standard = store.get_standard(standard_id)
        for conflict in standard.get("compiled", {}).get("conflicts", []):
            if conflict.get("kind") == "boundary" and not conflict.get("source_excerpts"):
                conflict["source_excerpts"] = _conflict_source_excerpts(conflict)
        compiled = parse_compiled_standard(standard["compiled"])
        return {
            **standard,
            "validation": validate_standard(compiled).model_dump(),
            "files": standard_to_yaml_files(compiled),
            "rule_stats": store.rule_stats(standard_id),
            "changes": store.list_standard_changes(standard_id),
        }
    except Exception as exc:
        raise _handle_error(exc) from exc


@app.delete("/api/standards/{standard_id}")
def delete_standard(standard_id: str) -> Dict[str, Any]:
    try:
        return store.delete_standard(standard_id)
    except Exception as exc:
        raise _handle_error(exc) from exc


@app.post("/api/standards/{standard_id}/activate")
def activate_standard(standard_id: str) -> Dict[str, Any]:
    try:
        standard = store.get_standard(standard_id)
        report = validate_standard(parse_compiled_standard(standard["compiled"]))
        if not report.valid:
            raise ValueError("Standard 校验未通过，不能激活")
        return store.activate_standard(standard_id)
    except Exception as exc:
        raise _handle_error(exc) from exc


@app.post("/api/standards/{standard_id}/revise")
def revise_standard(standard_id: str, payload: RuleRevisionRequest) -> Dict[str, Any]:
    try:
        result = service.revise_rule(
            standard_id,
            payload.rule_id,
            payload.new_rule,
            payload.reason,
            payload.related_case_ids,
        )
        if payload.suggestion_id:
            store.update_suggestion_status(payload.suggestion_id, "accepted")
        return result
    except Exception as exc:
        raise _handle_error(exc) from exc


@app.post("/api/standards/{standard_id}/versions")
def create_manual_standard_version(
    standard_id: str, payload: ManualVersionRequest
) -> Dict[str, Any]:
    try:
        return service.create_manual_version(
            standard_id,
            payload.compiled,
            payload.reason,
            payload.resolve_conflicts,
        )
    except Exception as exc:
        raise _handle_error(exc) from exc


def _parse_upload(filename: str, content: bytes) -> List[Dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".txt":
        text = _decode_upload_text(content)
        return [{"text": line.strip()} for line in text.splitlines() if line.strip()]
    if suffix == ".csv":
        text = _decode_upload_text(content)
        reader = csv.DictReader(io.StringIO(text))
        headers = [str(header or "").strip() for header in (reader.fieldnames or [])]
        if "text" not in headers:
            raise ValueError("CSV 必须包含 text 表头")
        return [
            {
                str(key or "").strip(): value
                for key, value in row.items()
                if str(key or "").strip()
            }
            for row in reader
        ]
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise ValueError("服务端未安装 XLSX 解析依赖 openpyxl") from exc
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        rows: List[Dict[str, Any]] = []
        for sheet in workbook.worksheets:
            values = list(sheet.iter_rows(values_only=True))
            if not values:
                continue
            headers = [str(value or "").strip() for value in values[0]]
            while headers and not headers[-1]:
                headers.pop()
            if "text" not in headers:
                raise ValueError(f"XLSX 工作表「{sheet.title}」必须包含 text 表头")
            for values_row in values[1:]:
                row = {
                    header: value
                    for header, value in zip(headers, values_row)
                    if header
                }
                if any(str(value or "").strip() for value in row.values()):
                    rows.append(row)
        return rows
    if suffix in {".jsonl", ".ndjson"}:
        text = _decode_upload_text(content)
        rows = []
        for line_number, line in enumerate(text.splitlines(), start=1):
            if not line.strip():
                continue
            try:
                value = json.loads(line)
            except json.JSONDecodeError as exc:
                raise ValueError(f"JSONL 第 {line_number} 行无效: {exc}") from exc
            if not isinstance(value, dict):
                raise ValueError(f"JSONL 第 {line_number} 行必须是对象")
            if "text" not in value:
                raise ValueError(f"JSONL 第 {line_number} 行缺少 text 字段")
            rows.append(value)
        return rows
    raise ValueError("仅支持 .csv、.xlsx、.txt、.jsonl 或 .ndjson 文件")


def _decode_upload_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("文本文件编码无法识别，请使用 UTF-8 或 GB18030")


@app.get("/api/datasets/template")
def dataset_template() -> Response:
    content = "text,gold_label\n示例文本,\n"
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="dataset-template.csv"'},
    )


@app.post("/api/datasets")
async def upload_dataset(
    file: UploadFile = File(...), name: Optional[str] = Form(default=None)
) -> Dict[str, Any]:
    try:
        content = await file.read()
        items = _parse_upload(file.filename or "", content)
        return store.create_dataset(name or Path(file.filename or "dataset").stem, file.filename, items)
    except Exception as exc:
        raise _handle_error(exc) from exc


@app.delete("/api/datasets/{dataset_id}")
def delete_dataset(dataset_id: str) -> Dict[str, Any]:
    try:
        return store.delete_dataset(dataset_id)
    except Exception as exc:
        raise _handle_error(exc) from exc


@app.get("/api/datasets")
def datasets() -> List[Dict[str, Any]]:
    return store.list_datasets()


@app.get("/api/datasets/{dataset_id}/items")
def dataset_items(dataset_id: str) -> List[Dict[str, Any]]:
    try:
        store.get_dataset(dataset_id)
        return store.list_items(dataset_id)
    except Exception as exc:
        raise _handle_error(exc) from exc


@app.post("/api/runs")
def create_run(payload: RunRequest, background_tasks: BackgroundTasks) -> Dict[str, Any]:
    try:
        standard = store.get_standard(payload.standard_id)
        if standard["status"] == "draft":
            raise ValueError("草稿 Standard 不能直接运行，请先激活")
        store.get_dataset(payload.dataset_id)
        run = store.create_run(
            payload.dataset_id, payload.standard_id, concurrency=payload.concurrency,
            trace_replicas=payload.trace_replicas,
        )
        background_tasks.add_task(service.process_run, run["id"])
        return run
    except Exception as exc:
        raise _handle_error(exc) from exc


@app.get("/api/runs")
def runs() -> List[Dict[str, Any]]:
    return store.list_runs()


@app.post("/api/runs/{run_id}/retry")
def retry_run(
    run_id: str,
    background_tasks: BackgroundTasks,
    concurrency: Optional[int] = None,
    trace_replicas: Optional[int] = None,
) -> Dict[str, Any]:
    try:
        run = store.get_run(run_id)
        if run["status"] != "failed":
            raise ValueError("只能继续失败的运行")
        # Continuing keeps the original parallelism unless the caller overrides it.
        updates: Dict[str, Any] = {"status": "queued", "error": None, "completed_at": None}
        if concurrency is not None:
            if not 1 <= concurrency <= MAX_RUN_CONCURRENCY:
                raise ValueError(f"并行度必须在 1 到 {MAX_RUN_CONCURRENCY} 之间")
            updates["concurrency"] = concurrency
        if trace_replicas is not None:
            if not 1 <= trace_replicas <= MAX_TRACE_REPLICAS:
                raise ValueError(f"Trace 副本数必须在 1 到 {MAX_TRACE_REPLICAS} 之间")
            updates["trace_replicas"] = trace_replicas
        store.update_run(run_id, **updates)
        background_tasks.add_task(service.process_run, run_id)
        return store.get_run(run_id)
    except Exception as exc:
        raise _handle_error(exc) from exc


@app.get("/api/runs/{run_id}")
def run_detail(run_id: str) -> Dict[str, Any]:
    try:
        return {
            "run": store.get_run(run_id),
            "metrics": store.run_metrics(run_id),
            "annotations": store.list_annotations(run_id),
            "events": store.list_annotation_events(run_id),
            "model_calls": store.list_model_calls(run_id),
        }
    except Exception as exc:
        raise _handle_error(exc) from exc


@app.get("/api/runs/{run_id}/events")
async def run_events(run_id: str, after: int = 0) -> StreamingResponse:
    """Stream durable query/model events; reconnects can resume from sequence."""
    try:
        store.get_run(run_id)
    except Exception as exc:
        raise _handle_error(exc) from exc

    async def stream():
        cursor = max(after, 0)
        while True:
            events = store.list_annotation_events(run_id, after=cursor)
            for event in events:
                cursor = event["sequence"]
                yield f"id: {cursor}\ndata: {json.dumps(event, ensure_ascii=False)}\n\n"
            try:
                run = store.get_run(run_id)
            except KeyError:
                return
            if run["status"] in {"completed", "failed"} and not events:
                return
            yield ": keep-alive\n\n"
            await asyncio.sleep(0.5)

    return StreamingResponse(
        stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@app.get("/api/runs/{run_id}/export")
def export_run(run_id: str, format: str = "jsonl", gold_only: bool = False) -> Response:
    try:
        annotations = store.list_annotations(run_id)
        if gold_only:
            annotations = [
                item for item in annotations
                if item["route"] == "AUTO_ACCEPT" or item.get("human_label")
            ]
        rows = [
            {
                "id": item["item_id"],
                "text": item["text"],
                "label": item.get("human_label") or item.get("label"),
                "decision_status": item["decision"]["status"],
                "decision_reason": item["decision"]["reason"],
                "route": item["route"],
                "route_reasons": item["route_reasons"],
                "rules_used": item["rules_used"],
                "evidence": item["evidence"],
                "confidence": item["confidence"],
                "gold_label": item.get("gold_label"),
            }
            for item in annotations
        ]
        suffix = "gold" if gold_only else "all"
        if format == "jsonl":
            content = "\n".join(json.dumps(row, ensure_ascii=False) for row in rows) + ("\n" if rows else "")
            return Response(
                content=content,
                media_type="application/x-ndjson",
                headers={"Content-Disposition": f'attachment; filename="labelspec-{suffix}.jsonl"'},
            )
        if format == "csv":
            stream = io.StringIO()
            fields = [
                "id", "text", "label", "decision_status", "decision_reason", "route",
                "route_reasons", "rules_used", "evidence", "confidence", "gold_label",
            ]
            writer = csv.DictWriter(stream, fieldnames=fields)
            writer.writeheader()
            for row in rows:
                writer.writerow({
                    **row,
                    "rules_used": "|".join(row["rules_used"]),
                    "route_reasons": "|".join(
                        reason["message"] for reason in row["route_reasons"]
                    ),
                })
            return Response(
                content="\ufeff" + stream.getvalue(),
                media_type="text/csv",
                headers={"Content-Disposition": f'attachment; filename="labelspec-{suffix}.csv"'},
            )
        raise ValueError("format 仅支持 csv 或 jsonl")
    except Exception as exc:
        raise _handle_error(exc) from exc


@app.post("/api/annotations/{annotation_id}/review")
def review(annotation_id: str, payload: ReviewRequest) -> Dict[str, Any]:
    try:
        return store.review_annotation(
            annotation_id,
            payload.human_label,
            payload.note,
            payload.reason_code,
            payload.rule_feedback,
        )
    except Exception as exc:
        raise _handle_error(exc) from exc


@app.get("/api/feedback")
def feedback(run_id: Optional[str] = None, standard_id: Optional[str] = None) -> List[Dict[str, Any]]:
    return store.list_feedback(run_id=run_id, standard_id=standard_id)


@app.post("/api/rule-patches")
def create_rule_patch(payload: RulePatchRequest) -> Dict[str, Any]:
    try:
        store.get_standard(payload.standard_id)
        return store.save_rule_patch(
            payload.standard_id,
            payload.payload,
            payload.related_feedback_ids,
            payload.source_run_id,
        )
    except Exception as exc:
        raise _handle_error(exc) from exc


@app.get("/api/rule-patches")
def rule_patches(standard_id: Optional[str] = None, status: Optional[str] = None) -> List[Dict[str, Any]]:
    return store.list_rule_patches(standard_id=standard_id, status=status)


@app.patch("/api/rule-patches/{patch_id}")
def update_rule_patch(
    patch_id: str,
    status: Optional[str] = None,
    payload: Optional[RulePatchEditRequest] = None,
) -> Dict[str, Any]:
    try:
        result = store.get_rule_patch(patch_id)
        if payload is not None:
            result = store.update_rule_patch_payload(patch_id, payload.payload)
        if status is None:
            return result
        if status == "applied":
            raise ValueError("Rule Patch 必须通过 apply 接口生效，不能直接标记 applied")
        return store.update_rule_patch_status(patch_id, status)
    except Exception as exc:
        raise _handle_error(exc) from exc


@app.post("/api/rule-patches/{patch_id}/apply")
def apply_rule_patch(patch_id: str, background_tasks: BackgroundTasks) -> Dict[str, Any]:
    try:
        result = service.apply_rule_patch(patch_id)
        # A patch creates an immutable successor. Activating it archives the
        # previous active version without deleting it, so historical v1 runs
        # and explicit reruns against v1 remain available.
        result["standard"] = store.activate_standard(result["standard"]["id"])
        patch = result["patch"]
        source_run_id = patch.get("source_run_id")
        operations = patch.get("payload", {}).get("operations", [])
        rule_id = next(
            (
                operation.get("rule_id") or operation.get("after", {}).get("rule_id")
                for operation in operations
                if operation.get("rule_id") or operation.get("after", {}).get("rule_id")
            ),
            "PATCH",
        )
        if source_run_id:
            impact = service.create_impact_run(
                source_run_id,
                result["standard"]["id"],
                rule_id,
                result.get("affected_labels", []),
            )
            background_tasks.add_task(service.process_run, impact["id"])
            result["impact_run"] = impact
        return result
    except Exception as exc:
        raise _handle_error(exc) from exc


@app.post("/api/runs/{run_id}/mine")
async def mine_run(run_id: str) -> Dict[str, Any]:
    try:
        run = store.get_run(run_id)
        if run["status"] != "completed":
            raise ValueError("只能分析已完成的运行")
        return {"suggestions": await miner.mine(run_id)}
    except Exception as exc:
        raise _handle_error(exc) from exc


@app.get("/api/suggestions")
def suggestions(run_id: Optional[str] = None) -> List[Dict[str, Any]]:
    return store.list_suggestions(run_id)


@app.patch("/api/suggestions/{suggestion_id}")
def update_suggestion(suggestion_id: str, status: str) -> Dict[str, Any]:
    try:
        return store.update_suggestion_status(suggestion_id, status)
    except Exception as exc:
        raise _handle_error(exc) from exc


@app.post("/api/impact-runs")
def impact_run(payload: ImpactRunRequest, background_tasks: BackgroundTasks) -> Dict[str, Any]:
    try:
        run = service.create_impact_run(
            payload.source_run_id,
            payload.target_standard_id,
            payload.rule_id,
            payload.labels,
            payload.concurrency,
            payload.trace_replicas,
        )
        background_tasks.add_task(service.process_run, run["id"])
        return run
    except Exception as exc:
        raise _handle_error(exc) from exc


@app.get("/api/compare")
def compare_runs(left_run_id: str, right_run_id: str) -> Dict[str, Any]:
    try:
        return {
            "left": {"run": store.get_run(left_run_id), "metrics": store.run_metrics(left_run_id)},
            "right": {"run": store.get_run(right_run_id), "metrics": store.run_metrics(right_run_id)},
        }
    except Exception as exc:
        raise _handle_error(exc) from exc


frontend_dist = Path(__file__).resolve().parents[2] / "frontend" / "dist"
if (frontend_dist / "index.html").exists():
    app.mount("/", StaticFiles(directory=frontend_dist, html=True), name="web")
